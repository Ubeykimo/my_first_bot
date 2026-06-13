from aiogram import Router, F
from aiogram.types import Message, CallbackQuery
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
import database as db
import keyboards as kb
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from aiogram.types import BufferedInputFile

router = Router()

# Состояния администратора
class AdminStates(StatesGroup):
    adding_service_name = State()
    adding_service_price = State()
    adding_service_duration = State()
    adding_schedule_start = State()  # убрали adding_schedule_day
    adding_schedule_end = State()
    editing_welcome = State()
    editing_info = State()

# Команда /admin
@router.message(F.text == "/admin")
async def admin_panel(message: Message, state: FSMContext):
    await state.clear()
    admin_id = await db.get_setting("admin_id")
    
    # Если админ не установлен — первый кто написал становится админом
    if not admin_id:
        await db.set_setting("admin_id", str(message.from_user.id))
        admin_id = str(message.from_user.id)
    
    if str(message.from_user.id) != admin_id:
        await message.answer("❌ У вас нет доступа")
        return
    
    await message.answer("👨‍💼 Панель администратора", reply_markup=kb.admin_menu())

# Меню админа
@router.callback_query(F.data == "admin_menu")
async def admin_menu(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.edit_text(
        "👨‍💼 Панель администратора",
        reply_markup=kb.admin_menu()
    )

# Все записи
@router.callback_query(F.data == "admin_bookings")
async def admin_bookings(callback: CallbackQuery):
    bookings = await db.get_bookings()
    if not bookings:
        await callback.message.edit_text(
            "Записей нет.",
            reply_markup=kb.admin_menu()
        )
        return
    
    text = "📋 Все записи:\n\n"
    for booking in bookings:
        text += (
            f"👤 {booking[1]}\n"
            f"✂️ {booking[2]}\n"
            f"📅 {booking[3]} в {booking[4]}\n"
            f"Статус: {booking[5]}\n\n"
        )
    
    await callback.message.edit_text(text, reply_markup=kb.admin_menu())

# Управление услугами
@router.callback_query(F.data == "admin_services")
async def admin_services(callback: CallbackQuery):
    services = await db.get_services()
    await callback.message.edit_text(
        "✂️ Управление услугами:",
        reply_markup=kb.admin_services_keyboard(services)
    )

# Удалить услугу
@router.callback_query(F.data.startswith("del_service_"))
async def delete_service(callback: CallbackQuery):
    service_id = int(callback.data.split("_")[2])
    await db.delete_service(service_id)
    services = await db.get_services()
    await callback.message.edit_text(
        "✂️ Услуга удалена. Управление услугами:",
        reply_markup=kb.admin_services_keyboard(services)
    )

# Добавить услугу — шаг 1
@router.callback_query(F.data == "add_service")
async def add_service_start(callback: CallbackQuery, state: FSMContext):
    await state.set_state(AdminStates.adding_service_name)
    await callback.message.edit_text("Введите название услуги:")

# Добавить услугу — шаг 2
@router.message(AdminStates.adding_service_name)
async def add_service_name(message: Message, state: FSMContext):
    await state.update_data(service_name=message.text)
    await state.set_state(AdminStates.adding_service_price)
    await message.answer("Введите цену услуги (только цифры):")

# Добавить услугу — шаг 3
@router.message(AdminStates.adding_service_price)
async def add_service_price(message: Message, state: FSMContext):
    if not message.text.isdigit():
        await message.answer("Введите только цифры!")
        return
    await state.update_data(service_price=int(message.text))
    await state.set_state(AdminStates.adding_service_duration)
    await message.answer("Введите длительность в минутах (например 30 или 60):")

# Добавить услугу — шаг 4
@router.message(AdminStates.adding_service_duration)
async def add_service_duration(message: Message, state: FSMContext):
    if not message.text.isdigit():
        await message.answer("Введите только цифры!")
        return
    data = await state.get_data()
    await db.add_service(data['service_name'], data['service_price'], int(message.text))
    await state.clear()
    services = await db.get_services()
    await message.answer(
        f"✅ Услуга '{data['service_name']}' добавлена!",
        reply_markup=kb.admin_services_keyboard(services)
    )

@router.callback_query(F.data == "admin_schedule")
async def admin_schedule(callback: CallbackQuery):
    schedule = await db.get_schedule()
    await callback.message.edit_text(
        "🕐 Управление расписанием:",
        reply_markup=kb.admin_schedule_keyboard(schedule)
    )

@router.callback_query(F.data.startswith("del_schedule_"))
async def delete_schedule(callback: CallbackQuery):
    schedule_id = int(callback.data.split("_")[2])
    await db.delete_schedule(schedule_id)
    schedule = await db.get_schedule()
    await callback.message.edit_text(
        "🕐 День удалён.",
        reply_markup=kb.admin_schedule_keyboard(schedule)
    )

@router.callback_query(F.data == "add_schedule")
async def add_schedule_start(callback: CallbackQuery, state: FSMContext):
    schedule = await db.get_schedule()
    existing_days = [slot[1] for slot in schedule]
    
    if len(existing_days) >= 7:
        await callback.answer("Все дни уже добавлены!", show_alert=True)
        return
    
    await callback.message.edit_text(
        "Выберите день недели:",
        reply_markup=kb.days_of_week_keyboard(existing_days)
    )

@router.callback_query(F.data == "ignore")
async def ignore(callback: CallbackQuery):
    await callback.answer("Этот день уже добавлен!", show_alert=True)

@router.callback_query(F.data.startswith("day_"))
async def schedule_day_chosen(callback: CallbackQuery, state: FSMContext):
    day = int(callback.data.split("_")[1])
    await state.update_data(schedule_day=day)
    await state.set_state(AdminStates.adding_schedule_start)
    await callback.message.edit_text(
        "Введите время начала работы:\n\n"
        "⚠️ Формат: ЧЧ:ММ, минуты кратные 15\n"
        "Например: 09:00, 09:15, 09:30, 09:45"
    )

@router.message(AdminStates.adding_schedule_start)
async def add_schedule_start_time(message: Message, state: FSMContext):
    # Проверяем формат
    try:
        from datetime import datetime
        time = datetime.strptime(message.text, "%H:%M")
        if time.minute % 15 != 0:
            await message.answer(
                "❌ Минуты должны быть кратны 15!\n"
                "Например: 09:00, 09:15, 09:30, 09:45"
            )
            return
    except ValueError:
        await message.answer(
            "❌ Неверный формат! Введите время в формате ЧЧ:ММ\n"
            "Например: 09:00"
        )
        return
    
    await state.update_data(schedule_start=message.text)
    await state.set_state(AdminStates.adding_schedule_end)
    await message.answer(
        "Введите время конца работы:\n\n"
        "⚠️ Формат: ЧЧ:ММ, минуты кратные 15\n"
        "Например: 18:00, 18:15, 18:30, 18:45"
    )

@router.message(AdminStates.adding_schedule_end)
async def add_schedule_end_time(message: Message, state: FSMContext):
    # Проверяем формат
    try:
        from datetime import datetime
        time = datetime.strptime(message.text, "%H:%M")
        if time.minute % 15 != 0:
            await message.answer(
                "❌ Минуты должны быть кратны 15!\n"
                "Например: 18:00, 18:15, 18:30, 18:45"
            )
            return
    except ValueError:
        await message.answer(
            "❌ Неверный формат! Введите время в формате ЧЧ:ММ\n"
            "Например: 18:00"
        )
        return
    
    data = await state.get_data()
    days = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
    
    # Проверяем что конец после начала
    from datetime import datetime
    start = datetime.strptime(data['schedule_start'], "%H:%M")
    end = datetime.strptime(message.text, "%H:%M")
    if end <= start:
        await message.answer("❌ Время конца должно быть позже времени начала!")
        return
    
    await db.add_schedule(data['schedule_day'], data['schedule_start'], message.text)
    await state.clear()
    schedule = await db.get_schedule()
    await message.answer(
        f"✅ Добавлено: {days[data['schedule_day']]} {data['schedule_start']}-{message.text}",
        reply_markup=kb.admin_schedule_keyboard(schedule)
    )

# Настройки
@router.callback_query(F.data == "admin_settings")
async def admin_settings(callback: CallbackQuery):
    welcome = await db.get_setting("welcome_text") or "Не задано"
    info_text = await db.get_setting("info_text") or "Не задано"
    await callback.message.edit_text(
        f"⚙️ Настройки\n\n"
        f"👋 Приветствие:\n{welcome}\n\n"
        f"ℹ️ Информация:\n{info_text}",
        reply_markup=kb.settings_keyboard()
    )

# Сохранить приветствие
@router.message(AdminStates.editing_welcome)
async def save_welcome(message: Message, state: FSMContext):
    await db.set_setting("welcome_text", message.text)
    await state.clear()
    await message.answer(
        "✅ Приветствие обновлено!",
        reply_markup=kb.settings_keyboard()
    )
    
# Подтверждение записи администратором
@router.callback_query(F.data.startswith("confirm_"))
async def confirm_booking_admin(callback: CallbackQuery):
    parts = callback.data.split("_")
    booking_id = int(parts[1])
    user_id = int(parts[2])
    
    booking = await db.get_booking_by_id(booking_id)
    if not booking:
        await callback.answer("Запись не найдена", show_alert=True)
        return
    
    await db.confirm_booking(booking_id)
    
    # Уведомление клиенту
    await callback.bot.send_message(
        user_id,
        f"✅ Ваша запись подтверждена!\n\n"
        f"✂️ Услуга: {booking[3]}\n"
        f"📅 Дата: {booking[4]}\n"
        f"🕐 Время: {booking[5]}\n\n"
        f"Ждём вас!"
    )
    
    await callback.message.edit_text(
        f"✅ Запись подтверждена!\n\n"
        f"👤 Клиент: {booking[2]}\n"
        f"✂️ Услуга: {booking[3]}\n"
        f"📅 {booking[4]} в {booking[5]}"
    )

# Отклонение записи администратором
@router.callback_query(F.data.startswith("reject_"))
async def reject_booking_admin(callback: CallbackQuery):
    parts = callback.data.split("_")
    booking_id = int(parts[1])
    user_id = int(parts[2])
    
    booking = await db.get_booking_by_id(booking_id)
    if not booking:
        await callback.answer("Запись не найдена", show_alert=True)
        return
    
    await db.reject_booking(booking_id)
    
    # Уведомление клиенту
    await callback.bot.send_message(
        user_id,
        f"❌ Ваша запись отклонена.\n\n"
        f"✂️ Услуга: {booking[3]}\n"
        f"📅 Дата: {booking[4]}\n"
        f"🕐 Время: {booking[5]}\n\n"
        f"Пожалуйста, выберите другое время."
    )
    
    await callback.message.edit_text(
        f"❌ Запись отклонена.\n\n"
        f"👤 Клиент: {booking[2]}\n"
        f"✂️ Услуга: {booking[3]}\n"
        f"📅 {booking[4]} в {booking[5]}"
    )
    
class AdminStates(StatesGroup):
    adding_service_name = State()
    adding_service_price = State()
    adding_service_duration = State()
    adding_schedule_day = State()
    adding_schedule_start = State()
    adding_schedule_end = State()
    editing_welcome = State()
    editing_info = State()  # новое
    
@router.callback_query(F.data == "edit_welcome")
async def edit_welcome(callback: CallbackQuery, state: FSMContext):
    await state.set_state(AdminStates.editing_welcome)
    await callback.message.edit_text("Введите новый текст приветствия:")

@router.callback_query(F.data == "edit_info")
async def edit_info(callback: CallbackQuery, state: FSMContext):
    await state.set_state(AdminStates.editing_info)
    await callback.message.edit_text("Введите текст для раздела Информация:")

@router.message(AdminStates.editing_info)
async def save_info(message: Message, state: FSMContext):
    await db.set_setting("info_text", message.text)
    await state.clear()
    await message.answer(
        "✅ Информация обновлена!",
        reply_markup=kb.settings_keyboard()
    )
    
@router.callback_query(F.data == "admin_stats")
async def admin_stats(callback: CallbackQuery):
    async with __import__('aiosqlite').connect("bot.db") as db_conn:
        # Всего записей
        cursor = await db_conn.execute("SELECT COUNT(*) FROM bookings")
        total = (await cursor.fetchone())[0]
        
        # Активных записей
        cursor = await db_conn.execute(
            "SELECT COUNT(*) FROM bookings WHERE status != 'cancelled'"
        )
        active = (await cursor.fetchone())[0]
        
        # Отменённых
        cursor = await db_conn.execute(
            "SELECT COUNT(*) FROM bookings WHERE status = 'cancelled'"
        )
        cancelled = (await cursor.fetchone())[0]
        
        # Подтверждённых
        cursor = await db_conn.execute(
            "SELECT COUNT(*) FROM bookings WHERE status = 'confirmed'"
        )
        confirmed = (await cursor.fetchone())[0]
        
        # Записей сегодня
        from datetime import datetime
        today = datetime.now().strftime("%d.%m.%Y")
        cursor = await db_conn.execute(
            "SELECT COUNT(*) FROM bookings WHERE date = ? AND status != 'cancelled'",
            (today,)
        )
        today_count = (await cursor.fetchone())[0]
    
    await callback.message.edit_text(
        f"📊 Статистика\n\n"
        f"📋 Всего записей: {total}\n"
        f"✅ Подтверждённых: {confirmed}\n"
        f"⏳ Активных: {active}\n"
        f"❌ Отменённых: {cancelled}\n"
        f"📅 Сегодня: {today_count}",
        reply_markup=kb.admin_menu()
    )
    
@router.callback_query(F.data == "admin_export")
async def admin_export(callback: CallbackQuery):
    await callback.message.edit_text(
        "📥 Выберите формат экспорта:",
        reply_markup=kb.export_keyboard()
    )

@router.callback_query(F.data == "export_excel")
async def export_excel(callback: CallbackQuery):
    bookings = await db.get_bookings()
    
    if not bookings:
        await callback.answer("Записей нет!", show_alert=True)
        return
    
    # Создаём Excel файл
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Записи"
    
    # Заголовки
    headers = ["№", "Клиент", "Услуга", "Дата", "Время", "Статус"]
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Данные
    status_map = {
        "pending": "⏳ Ожидает",
        "confirmed": "✅ Подтверждена",
        "cancelled": "❌ Отменена"
    }
    
    for row, booking in enumerate(bookings, 2):
        ws.cell(row=row, column=1, value=booking[0])
        ws.cell(row=row, column=2, value=booking[1])
        ws.cell(row=row, column=3, value=booking[2])
        ws.cell(row=row, column=4, value=booking[3])
        ws.cell(row=row, column=5, value=booking[4])
        ws.cell(row=row, column=6, value=status_map.get(booking[5], booking[5]))
    
    # Ширина колонок
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 18
    
    # Сохраняем в память
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    from datetime import datetime
    filename = f"записи_{datetime.now().strftime('%d.%m.%Y')}.xlsx"
    
    await callback.message.answer_document(
        BufferedInputFile(buffer.read(), filename=filename),
        caption=f"📊 Экспорт записей\nВсего: {len(bookings)} записей"
    )
    await callback.answer()

@router.callback_query(F.data == "export_txt")
async def export_txt(callback: CallbackQuery):
    bookings = await db.get_bookings()
    
    if not bookings:
        await callback.answer("Записей нет!", show_alert=True)
        return
    
    status_map = {
        "pending": "⏳ Ожидает",
        "confirmed": "✅ Подтверждена",
        "cancelled": "❌ Отменена"
    }
    
    from datetime import datetime
    text = f"📋 Все записи — {datetime.now().strftime('%d.%m.%Y')}\n"
    text += "=" * 40 + "\n\n"
    
    for booking in bookings:
        text += (
            f"👤 {booking[1]}\n"
            f"✂️ {booking[2]}\n"
            f"📅 {booking[3]} в {booking[4]}\n"
            f"Статус: {status_map.get(booking[5], booking[5])}\n"
            f"{'—' * 30}\n"
        )
    
    filename = f"записи_{datetime.now().strftime('%d.%m.%Y')}.txt"
    
    await callback.message.answer_document(
        BufferedInputFile(text.encode('utf-8'), filename=filename),
        caption=f"📝 Экспорт записей\nВсего: {len(bookings)} записей"
    )
    await callback.answer()
    
# Завершить визит
@router.callback_query(F.data == "admin_complete")
async def admin_complete(callback: CallbackQuery):
    bookings = await db.get_completed_bookings()
    if not bookings:
        await callback.message.edit_text(
            "Нет подтверждённых записей.",
            reply_markup=kb.admin_menu()
        )
        return
    await callback.message.edit_text(
        "Выберите запись для завершения:",
        reply_markup=kb.completed_bookings_keyboard(bookings)
    )

@router.callback_query(F.data.startswith("complete_"))
async def complete_visit(callback: CallbackQuery):
    parts = callback.data.split("_")
    booking_id = int(parts[1])
    user_id = int(parts[2])
    
    booking = await db.get_booking_by_id(booking_id)
    if not booking:
        await callback.answer("Запись не найдена", show_alert=True)
        return
    
    await db.complete_booking(booking_id)
    
    # Отправляем клиенту запрос отзыва
    await callback.bot.send_message(
        user_id,
        f"Спасибо за визит! 😊\n\n"
        f"Вы посетили: {booking[3]}\n"
        f"Пожалуйста оцените услугу:",
        reply_markup=kb.rating_keyboard(booking_id)
    )
    
    await callback.message.edit_text(
        f"✅ Визит завершён!\n"
        f"Клиенту отправлен запрос отзыва.",
        reply_markup=kb.admin_menu()
    )

# Отзывы
@router.callback_query(F.data == "admin_reviews")
async def admin_reviews(callback: CallbackQuery):
    reviews = await db.get_reviews()
    text = kb.reviews_keyboard(reviews)
    await callback.message.edit_text(
        text,
        reply_markup=kb.admin_menu()
    )